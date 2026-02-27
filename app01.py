import os
import re
import csv
import openpyxl
import win32com.client as win32
from datetime import datetime
from email.utils import parseaddr

# ========== CONFIG ==========
deal_folder = r"\\10.242.132.6\NCR-Common6\ILD Reports\Akash Tripathi GB\Voice Deal"
image_path = r"C:\Users\b0310605\Desktop\Deal_Config\dashboard.png"
log_file = r"C:\Users\b0310605\Desktop\LOG\Deal_Auto_Responder_delegated_access_log.csv"
fixed_from = "ABC@airtel.com"
logo_path = r"C:\Users\b0310605\Desktop\Deal_Config\airtel.png"  # 👈 Airtel logo path

# Subject pattern e.g. Airtel_Deal Status: Carrier Name
subject_pattern = re.compile(r"^Airtel_Deal\s*Status\s*[\-–—:]\s*(.+)", re.I)

# ========== ACCESS CONTROL ==========
delegated_access = {
    "ABD@airtel.com": ["sameer hajelay", "vanshika aggarwal"],
    "XYZl@airtel.com": ["*"],
    "anand@airtel.com": ["*"],
    "ashish1.shukla@airtel.com": ["*"],
    "akash.tripathi@airtel.com": ["*"],
    "rahul5.sharma@airtel.com": ["*"],
    "Anukrati3.Rajvanshi@airtel.com": ["*"],
}
delegated_access = {k.lower(): [am.lower() for am in v] for k, v in delegated_access.items()}

# ========== LOG FILE SETUP ==========
os.makedirs(os.path.dirname(log_file), exist_ok=True)
if not os.path.exists(log_file):
    with open(log_file, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerow(["Timestamp", "Sender", "Carrier", "Result", "Reason", "ExcelFile"])


def log_event(sender, carrier, result, reason, excel_file):
    with open(log_file, "a", newline="", encoding="utf-8") as f:
        csv.writer(f).writerow(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                sender,
                carrier,
                result,
                reason,
                excel_file or "N/A",
            ]
        )


def normalize(text):
    if not text:
        return ""
    return str(text).replace("\xa0", " ").replace("\u200b", "").strip().lower()


def is_airtel_sender(sender_email):
    return (sender_email or "").strip().lower().endswith("@airtel.com")


# ========== FIND CARRIER FILE ==========
def find_carrier_file(carrier_name):
    target = normalize(carrier_name)
    for root, _, files in os.walk(deal_folder):
        for f in files:
            if f.endswith(".xlsx") and not f.startswith("~$"):
                path = os.path.join(root, f)
                try:
                    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
                    if "dashboard" in wb.sheetnames:
                        ws = wb["dashboard"]
                        c2 = normalize(ws["C2"].value)
                        am = (ws["C5"].value or "").strip()
                        if c2 == target:
                            wb.close()
                            return path, am
                    wb.close()
                except Exception:
                    continue
    return None, None


# ========== EXPORT DASHBOARD IMAGE ==========
def export_dashboard_image(file_path):
    os.system("taskkill /f /im excel.exe >nul 2>&1")
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(file_path, ReadOnly=True)
    ws = wb.Sheets("dashboard")
    used_range = ws.UsedRange
    used_range.CopyPicture(Format=2)
    chart = ws.ChartObjects().Add(0, 0, used_range.Width + 10, used_range.Height + 10)
    chart.Chart.Paste()
    chart.Chart.Export(image_path)
    chart.Delete()
    wb.Close(False)
    excel.Quit()
    return image_path


# ========== REPLY SUCCESS MAIL ==========
def reply_with_status(original_msg, carrier, excel_file):
    outlook = win32.Dispatch("Outlook.Application")
    reply = original_msg.Reply()
    reply.Subject = f"Re: Airtel_Deal Status – {carrier}"

    # Use fixed From account
    reply.SendUsingAccount = [
        acc for acc in outlook.Session.Accounts if acc.SmtpAddress.lower() == fixed_from.lower()
    ][0]

    # Dashboard image
    img = reply.Attachments.Add(image_path)
    cid = "DashboardIMG"
    img.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", cid)

    # Airtel logo image
    logo = reply.Attachments.Add(logo_path)
    logo.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "airtelLogo")

    # Excel file
    reply.Attachments.Add(excel_file)

    # Email body with embedded images
    reply.HTMLBody = f"""
    <html>
    <body style="font-family: Calibri; font-size: 11pt; color: #000000;">
        Dear {original_msg.SenderName},<br><br>
        Please find below the current deal status for <b>{carrier}</b>.<br><br>
        <img src="cid:{cid}"><br><br>
        The complete Excel dashboard file is attached.<br><br>

        Regards,<br>
        <b>Airtel Deal Automation System</b><br>
        <img src="cid:airtelLogo" width="120" style="margin-top:6px; height:auto; border:0; display:inline-block;"><br>
        <span style="font-size:10pt; color:gray;">
            (This is an auto-generated response by the Airtel Deal Automation System.<br>
            Access is restricted to authorized @airtel.com users.<br>
            All requests are logged for security and audit purposes.)
        </span>
    </body>
    </html>
    """
    reply.Send()
    print(f"📤 Replied with deal status for {carrier}")


# ========== REPLY FAILURE MAIL ==========
def reply_with_failure(original_msg, reason, carrier=None):
    outlook = win32.Dispatch("Outlook.Application")
    reply = original_msg.Reply()
    reply.Subject = (
        f"Re: Airtel_Deal Status – {carrier} (Failed)"
        if carrier
        else "Re: Airtel_Deal Status (Failed)"
    )

    # Use fixed From account
    reply.SendUsingAccount = [
        acc for acc in outlook.Session.Accounts if acc.SmtpAddress.lower() == fixed_from.lower()
    ][0]

    # Airtel logo image
    logo = reply.Attachments.Add(logo_path)
    logo.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "airtelLogo")

    reply.HTMLBody = f"""
    <html>
    <body style="font-family: Calibri; font-size: 11pt; color: #000000;">
        Dear {original_msg.SenderName},<br><br>
        We could not process your Airtel deal status request.<br><br>
        <b>Reason:</b> {reason}<br><br>
        Please verify the subject format and access authorization, then resend your request.<br>
        Expected subject example: <b>Airtel_Deal Status - Vodafone</b><br><br>

        Regards,<br>
        <b>Airtel Deal Automation System</b><br>
        <img src="cid:airtelLogo" width="120" style="margin-top:6px; height:auto; border:0; display:inline-block;"><br>
        <span style="font-size:10pt; color:gray;">
            (This is an auto-generated response by the Airtel Deal Automation System.<br>
            Access is restricted to authorized @airtel.com users.<br>
            All requests are logged for security and audit purposes.)
        </span>
    </body>
    </html>
    """
    reply.Send()
    print(f"📤 Sent failure notification for {carrier or 'request'}")


# ========== GET SENDER EMAIL ==========
def get_smtp_address(msg):
    try:
        sender = msg.Sender
        if sender and sender.AddressEntryUserType == 0:
            return sender.GetExchangeUser().PrimarySmtpAddress.lower()
        _, addr = parseaddr(msg.SenderEmailAddress or "")
        return addr.lower()
    except Exception:
        _, addr = parseaddr(msg.SenderEmailAddress or "")
        return addr.lower()


# ========== PROCESS FOLDER ==========
def process_deal_status_folder(folder):
    messages = folder.Items
    messages.Sort("[ReceivedTime]", True)
    print(f"📂 Reading unread mails from folder: {folder.Name}")

    for i in range(1, messages.Count + 1):
        try:
            msg = messages.Item(i)
            if not getattr(msg, "Unread", False):
                continue

            sender_email = get_smtp_address(msg)
            subj = (msg.Subject or "").strip()
            m = subject_pattern.search(subj)
            if not m:
                reason = "Invalid subject format. Use: Airtel_Deal Status - <Carrier Name>."
                log_event(sender_email, "UNKNOWN", "FAILED", f"{reason} Subject: {subj}", None)
                print(f"❌ {reason} Subject: {subj}")
                if is_airtel_sender(sender_email):
                    reply_with_failure(msg, reason)
                msg.Unread = False
                continue

            carrier = m.group(1).strip()
            sender_name = (msg.SenderName or "").strip()

            print(f"📩 From {sender_email} requesting carrier '{carrier}'")

            # Find matching Excel + AM
            file_path, acct_mgr = find_carrier_file(carrier)
            if not file_path:
                reason = f"Carrier not found: {carrier}"
                print(f"❌ No Excel found for {carrier}")
                log_event(sender_email, carrier, "FAILED", reason, None)
                if is_airtel_sender(sender_email):
                    reply_with_failure(msg, reason, carrier)
                msg.Unread = False
                continue

            # ===== Authorization Check =====
            acct_norm = normalize(acct_mgr)
            sender_norm = normalize(sender_name)
            sender_local = sender_email.split("@")[0]
            sender_email_norm = sender_email.lower()

            authorized = False
            access_reason = ""

            if acct_norm in sender_norm or acct_norm == sender_local:
                authorized = True
                access_reason = "Own Deal Access"
            elif sender_email_norm in delegated_access:
                allowed_list = delegated_access[sender_email_norm]
                if "*" in allowed_list:
                    authorized = True
                    access_reason = "Delegated Full Access"
                else:
                    for am_name in allowed_list:
                        if am_name.lower() in acct_norm:
                            authorized = True
                            access_reason = f"Delegated via {sender_email_norm} → {acct_mgr}"
                            break

            if not authorized:
                reason = "Unauthorized access for requested deal"
                print(f"⛔ Sender '{sender_name}' not authorized (AM: {acct_mgr})")
                log_event(sender_email, carrier, "DENIED", reason, file_path)
                if is_airtel_sender(sender_email):
                    reply_with_failure(msg, reason, carrier)
                msg.Unread = False
                continue

            export_dashboard_image(file_path)
            reply_with_status(msg, carrier, file_path)
            log_event(sender_email, carrier, "SUCCESS", f"Replied ({access_reason})", file_path)
            msg.Unread = False

        except Exception as e:
            print("⚠️ Error processing mail:", e)
            log_event("UNKNOWN", "UNKNOWN", "ERROR", str(e), None)


# ========== MAIN ==========
def run():
    outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
    root = outlook.GetDefaultFolder(6).Parent

    deal_folder_obj = None
    for f in root.Folders:
        if f.Name.lower() == "deal status":
            deal_folder_obj = f
            break

    if not deal_folder_obj:
        print("Folder 'Deal Status' not found in Outlook.")
        return

    process_deal_status_folder(deal_folder_obj)
    print("Completed at", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


# ========== EXECUTION ==========
if __name__ == "__main__":
    print("Running Airtel Deal Auto-Responder (reading unread mails from 'Deal Status')...")
    run()
